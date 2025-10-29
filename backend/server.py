from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, UploadFile, File
from fastapi.responses import StreamingResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional
import uuid
from datetime import datetime, timezone, timedelta
import bcrypt
import jwt
import csv
import io
import json
import pandas as pd
import xmltodict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import secrets


ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# JWT Configuration
JWT_SECRET = os.environ.get('JWT_SECRET', 'your-secret-key-change-in-production')
JWT_ALGORITHM = 'HS256'
JWT_EXPIRATION_HOURS = 24

# Create the main app without a prefix
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

security = HTTPBearer()

# Define Models
class UserRegister(BaseModel):
    master_username: str
    master_password: str

class UserLogin(BaseModel):
    master_username: str
    master_password: str

class RecoveryRequest(BaseModel):
    master_username: str
    recovery_key: str
    new_master_password: str

class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    master_username: str
    master_password_hash: str
    salt: str
    recovery_key_hash: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class PasswordEntryCreate(BaseModel):
    title: str
    email: Optional[str] = None
    username: Optional[str] = None
    encrypted_password: str
    url: Optional[str] = None
    notes: Optional[str] = None

class PasswordEntryUpdate(BaseModel):
    title: Optional[str] = None
    email: Optional[str] = None
    username: Optional[str] = None
    encrypted_password: Optional[str] = None
    url: Optional[str] = None
    notes: Optional[str] = None

class PasswordEntry(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    title: str
    email: Optional[str] = None
    username: Optional[str] = None
    encrypted_password: str
    url: Optional[str] = None
    notes: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class TokenResponse(BaseModel):
    token: str
    user_id: str
    master_username: str
    recovery_key: Optional[str] = None

# Helper functions
def generate_recovery_key() -> str:
    """Generate a human-readable recovery key"""
    parts = [secrets.token_hex(4).upper() for _ in range(4)]
    return '-'.join(parts)

def hash_password(password: str, salt: bytes) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), salt).decode('utf-8')

def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))

def create_jwt_token(user_id: str, master_username: str) -> str:
    expiration = datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION_HOURS)
    payload = {
        'user_id': user_id,
        'master_username': master_username,
        'exp': expiration
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        token = credentials.credentials
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        return payload
    except jwt.ExpiredSignatureError:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Token expired"
        )
    except jwt.InvalidTokenError:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Invalid token"
        )

# Auth Routes
@api_router.post("/auth/register", response_model=TokenResponse)
async def register(user_data: UserRegister):
    # Check if user already exists
    existing_user = await db.users.find_one({"master_username": user_data.master_username}, {"_id": 0})
    if existing_user:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Username already exists"
        )
    
    # Generate recovery key and salt
    recovery_key = generate_recovery_key()
    salt = bcrypt.gensalt()
    recovery_salt = bcrypt.gensalt()
    
    # Hash passwords
    password_hash = hash_password(user_data.master_password, salt)
    recovery_key_hash = hash_password(recovery_key, recovery_salt)
    
    # Create user
    user = User(
        master_username=user_data.master_username,
        master_password_hash=password_hash,
        salt=salt.decode('utf-8'),
        recovery_key_hash=recovery_key_hash
    )
    
    # Save to database
    doc = user.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.users.insert_one(doc)
    
    # Create JWT token
    token = create_jwt_token(user.id, user.master_username)
    
    return TokenResponse(
        token=token,
        user_id=user.id,
        master_username=user.master_username,
        recovery_key=recovery_key  # Return only during registration
    )

@api_router.post("/auth/login", response_model=TokenResponse)
async def login(user_data: UserLogin):
    # Find user
    user = await db.users.find_one({"master_username": user_data.master_username}, {"_id": 0})
    if not user:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Invalid username or password"
        )
    
    # Verify password
    if not verify_password(user_data.master_password, user['master_password_hash']):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Invalid username or password"
        )
    
    # Create JWT token
    token = create_jwt_token(user['id'], user['master_username'])
    
    return TokenResponse(
        token=token,
        user_id=user['id'],
        master_username=user['master_username']
    )

@api_router.post("/auth/recover")
async def recover_password(recovery_data: RecoveryRequest):
    # Find user
    user = await db.users.find_one({"master_username": recovery_data.master_username}, {"_id": 0})
    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="User not found"
        )
    
    # Verify recovery key
    if not verify_password(recovery_data.recovery_key, user['recovery_key_hash']):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Invalid recovery key"
        )
    
    # Generate new password hash
    salt = bcrypt.gensalt()
    new_password_hash = hash_password(recovery_data.new_master_password, salt)
    
    # Update user password
    await db.users.update_one(
        {"master_username": recovery_data.master_username},
        {"$set": {
            "master_password_hash": new_password_hash,
            "salt": salt.decode('utf-8')
        }}
    )
    
    return {"message": "Password reset successfully"}

# Password Routes
@api_router.get("/passwords", response_model=List[PasswordEntry])
async def get_passwords(current_user: dict = Depends(get_current_user)):
    user_id = current_user['user_id']
    passwords = await db.password_entries.find({"user_id": user_id}, {"_id": 0}).to_list(1000)
    
    # Convert ISO string timestamps back to datetime objects
    for pwd in passwords:
        if isinstance(pwd.get('created_at'), str):
            pwd['created_at'] = datetime.fromisoformat(pwd['created_at'])
        if isinstance(pwd.get('updated_at'), str):
            pwd['updated_at'] = datetime.fromisoformat(pwd['updated_at'])
    
    return passwords

@api_router.post("/passwords", response_model=PasswordEntry)
async def create_password(password_data: PasswordEntryCreate, current_user: dict = Depends(get_current_user)):
    user_id = current_user['user_id']
    
    password_entry = PasswordEntry(
        user_id=user_id,
        **password_data.model_dump()
    )
    
    # Save to database
    doc = password_entry.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    doc['updated_at'] = doc['updated_at'].isoformat()
    await db.password_entries.insert_one(doc)
    
    return password_entry

@api_router.put("/passwords/{password_id}", response_model=PasswordEntry)
async def update_password(
    password_id: str,
    password_data: PasswordEntryUpdate,
    current_user: dict = Depends(get_current_user)
):
    user_id = current_user['user_id']
    
    # Find existing password
    existing = await db.password_entries.find_one({"id": password_id, "user_id": user_id}, {"_id": 0})
    if not existing:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Password entry not found"
        )
    
    # Update fields
    update_data = password_data.model_dump(exclude_unset=True)
    if update_data:
        update_data['updated_at'] = datetime.now(timezone.utc).isoformat()
        await db.password_entries.update_one(
            {"id": password_id, "user_id": user_id},
            {"$set": update_data}
        )
    
    # Get updated entry
    updated = await db.password_entries.find_one({"id": password_id}, {"_id": 0})
    if isinstance(updated.get('created_at'), str):
        updated['created_at'] = datetime.fromisoformat(updated['created_at'])
    if isinstance(updated.get('updated_at'), str):
        updated['updated_at'] = datetime.fromisoformat(updated['updated_at'])
    
    return PasswordEntry(**updated)

@api_router.delete("/passwords/{password_id}")
async def delete_password(password_id: str, current_user: dict = Depends(get_current_user)):
    user_id = current_user['user_id']
    
    result = await db.password_entries.delete_one({"id": password_id, "user_id": user_id})
    if result.deleted_count == 0:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Password entry not found"
        )
    
    return {"message": "Password deleted successfully"}

# Import/Export Routes
@api_router.post("/passwords/import")
async def import_passwords(file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    user_id = current_user['user_id']
    
    try:
        content = await file.read()
        file_extension = file.filename.split('.')[-1].lower()
        
        imported_count = 0
        
        if file_extension == 'csv':
            # Parse CSV
            df = pd.read_csv(io.BytesIO(content))
            imported_count = await process_import_dataframe(df, user_id)
            
        elif file_extension in ['xlsx', 'xlsm']:
            # Parse Excel
            df = pd.read_excel(io.BytesIO(content))
            imported_count = await process_import_dataframe(df, user_id)
            
        elif file_extension == 'xml':
            # Parse XML
            xml_dict = xmltodict.parse(content)
            entries = []
            
            # Handle different XML structures
            if 'passwords' in xml_dict and 'entry' in xml_dict['passwords']:
                entries = xml_dict['passwords']['entry']
                if not isinstance(entries, list):
                    entries = [entries]
            
            for entry in entries:
                password_entry = PasswordEntry(
                    user_id=user_id,
                    title=entry.get('title', entry.get('name', 'Untitled')),
                    email=entry.get('email'),
                    username=entry.get('username'),
                    encrypted_password=entry.get('encrypted_password', entry.get('password', '')),
                    url=entry.get('url'),
                    notes=entry.get('notes', entry.get('extra'))
                )
                
                doc = password_entry.model_dump()
                doc['created_at'] = doc['created_at'].isoformat()
                doc['updated_at'] = doc['updated_at'].isoformat()
                await db.password_entries.insert_one(doc)
                imported_count += 1
        else:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Unsupported file format"
            )
        
        return {"message": f"Successfully imported {imported_count} passwords"}
        
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Error importing file: {str(e)}"
        )

async def process_import_dataframe(df: pd.DataFrame, user_id: str) -> int:
    """Process DataFrame from CSV/Excel and import passwords"""
    imported_count = 0
    
    # Normalize column names
    df.columns = df.columns.str.lower().str.strip()
    
    # Map common column names from different password managers
    column_mapping = {
        'name': 'title',
        'website': 'url',
        'extra': 'notes',
        'password': 'encrypted_password',
        'grouping': 'notes'  # Add folder info to notes
    }
    
    df = df.rename(columns=column_mapping)
    
    # Process each row
    for _, row in df.iterrows():
        # Skip empty rows
        if pd.isna(row.get('title')) and pd.isna(row.get('url')):
            continue
        
        password_entry = PasswordEntry(
            user_id=user_id,
            title=str(row.get('title', 'Untitled')) if pd.notna(row.get('title')) else 'Untitled',
            email=str(row.get('email')) if pd.notna(row.get('email')) else None,
            username=str(row.get('username')) if pd.notna(row.get('username')) else None,
            encrypted_password=str(row.get('encrypted_password', '')) if pd.notna(row.get('encrypted_password')) else '',
            url=str(row.get('url')) if pd.notna(row.get('url')) else None,
            notes=str(row.get('notes')) if pd.notna(row.get('notes')) else None
        )
        
        doc = password_entry.model_dump()
        doc['created_at'] = doc['created_at'].isoformat()
        doc['updated_at'] = doc['updated_at'].isoformat()
        await db.password_entries.insert_one(doc)
        imported_count += 1
    
    return imported_count

@api_router.get("/passwords/export")
async def export_passwords(format: str = 'csv', current_user: dict = Depends(get_current_user)):
    user_id = current_user['user_id']
    
    # Get all passwords
    passwords = await db.password_entries.find({"user_id": user_id}, {"_id": 0}).to_list(1000)
    
    if not passwords:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="No passwords to export"
        )
    
    # Prepare data
    export_data = []
    for pwd in passwords:
        export_data.append({
            'title': pwd.get('title', ''),
            'email': pwd.get('email', ''),
            'username': pwd.get('username', ''),
            'encrypted_password': pwd.get('encrypted_password', ''),
            'url': pwd.get('url', ''),
            'notes': pwd.get('notes', '')
        })
    
    if format == 'csv':
        # Export as CSV
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=['title', 'email', 'username', 'encrypted_password', 'url', 'notes'])
        writer.writeheader()
        writer.writerows(export_data)
        
        return StreamingResponse(
            io.BytesIO(output.getvalue().encode('utf-8')),
            media_type='text/csv',
            headers={'Content-Disposition': 'attachment; filename=safepass_export.csv'}
        )
    
    elif format in ['xlsx', 'xlsm']:
        # Export as Excel
        wb = Workbook()
        ws = wb.active
        ws.title = 'Passwords'
        
        # Add headers with styling
        headers = ['Title', 'Email', 'Username', 'Encrypted Password', 'URL', 'Notes']
        ws.append(headers)
        
        # Style header row
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        # Add data
        for entry in export_data:
            ws.append([
                entry['title'],
                entry['email'],
                entry['username'],
                entry['encrypted_password'],
                entry['url'],
                entry['notes']
            ])
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        media_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' if format == 'xlsx' else 'application/vnd.ms-excel.sheet.macroEnabled.12'
        
        return StreamingResponse(
            output,
            media_type=media_type,
            headers={'Content-Disposition': f'attachment; filename=safepass_export.{format}'}
        )
    
    elif format == 'xml':
        # Export as XML
        xml_data = {
            'passwords': {
                'entry': [
                    {
                        'title': entry['title'],
                        'email': entry['email'],
                        'username': entry['username'],
                        'encrypted_password': entry['encrypted_password'],
                        'url': entry['url'],
                        'notes': entry['notes']
                    }
                    for entry in export_data
                ]
            }
        }
        
        xml_string = xmltodict.unparse(xml_data, pretty=True)
        
        return StreamingResponse(
            io.BytesIO(xml_string.encode('utf-8')),
            media_type='application/xml',
            headers={'Content-Disposition': 'attachment; filename=safepass_export.xml'}
        )
    
    else:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Unsupported export format"
        )

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()